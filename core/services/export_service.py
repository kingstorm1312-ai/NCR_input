import io
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import streamlit as st
import os

def generate_dnxl_docx(ncr_data, dnxl_data, details_df):
    """
    Điền dữ liệu DNXL vào template Excel (XLSX) sử dụng openpyxl.
    
    Args:
        ncr_data: dict info của NCR gốc
        dnxl_data: dict info của DNXL Master
        details_df: DataFrame chi tiết lỗi
        
    Returns:
        io.BytesIO: Buffer file XLSX
    """
    # Template Path (Excel) - Relative to project root
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = os.path.join(project_root, "templates", "Template_DNXL.xlsx")
    
    if not os.path.exists(template_path):
        st.error(f"❌ Không tìm thấy file mẫu: {template_path}")
        st.info("💡 Đảm bảo folder `templates/` và file `Template_DNXL.xlsx` tồn tại trong project.")
        return None

    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active # Assume first sheet
        
        # --- 1. PREPARE DATA ---
        total_assigned = 0
        total_fixed = 0
        if not details_df.empty:
            for _, r in details_df.iterrows():
                total_assigned += float(r.get('qty_assigned', 0) or 0)
                total_fixed += float(r.get('qty_fixed', 0) or 0)

        # Helper for date format
        def format_date_vn(d_str):
            if not d_str: return ""
            try:
                if isinstance(d_str, (pd.Timestamp, pd.DatetimeIndex)):
                     return d_str.strftime("%d/%m/%Y")
                
                # Check if it has time component
                if " " in str(d_str):
                    return pd.to_datetime(d_str).strftime("%d/%m/%Y")
                return str(d_str)
            except:
                return str(d_str)

        import re
        from openpyxl.cell.cell import MergedCell

        # --- 2. REPLACE PLACEHOLDERS (REGEX) ---
        # Map keys without brackets for regex matching
        # Data dictionary
        data_map = {
            'hop_dong': ncr_data.get('hop_dong', ''),
            'ten_sp': ncr_data.get('ten_sp', ''),
            'ma_vat_tu': ncr_data.get('ma_vat_tu', ''),
            'ncr_id': ncr_data.get('so_phieu_ncr') or ncr_data.get('so_phieu', ''),
            'ngay_ncr': format_date_vn(ncr_data.get('ngay_lap', '')),
            'nguoi_lap': dnxl_data.get('created_by', ''),
            'noi_sx': ncr_data.get('nguon_goc', ''),
            'noi_gay_loi': ncr_data.get('vi_tri_loi', ''),
            # User Change: sl_yeu_cau is manually input in Target Scope field
            'sl_yeu_cau': dnxl_data.get('target_scope', ''), 
            'ngay_dua_huong_xl': format_date_vn(dnxl_data.get('created_at', '')),
            'thoi_gian_xl': format_date_vn(dnxl_data.get('deadline', '')),
            'tong_sl_dat': total_fixed,
            'dnxl_id': dnxl_data.get('dnxl_id', ''),
            'target_scope': dnxl_data.get('target_scope', ''),
            'handling_instruction': dnxl_data.get('handling_instruction', '')
        }

        # Search first 50 rows, 20 cols
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
            for cell in row:
                # Skip if cell is part of a merge (but not top-left)
                if isinstance(cell, MergedCell):
                    continue
                    
                if cell.value and isinstance(cell.value, str):
                    val = str(cell.value)
                    
                    # Regex to find {{ key }} or {{key}} with any internal spacing
                    # Pattern: {{ \s*KEY \s* }}
                    for key, replace_val in data_map.items():
                        # Escape key just in case, though these are simple strings
                        pattern = r"\{\{\s*" + re.escape(key) + r"\s*\}\}"
                        if re.search(pattern, val):
                            val = re.sub(pattern, str(replace_val), val)
                            
                    cell.value = val

        # --- 3. WRITE TABLE DETAILS ---
        # Strategy: Find explicit template loop row {{ i.ten_loi }}
        start_row = None
        
        # 1. Search for value containing 'ten_loi' (The template tag)
        for row in ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "ten_loi" in cell.value and "{{" in cell.value:
                        start_row = cell.row
                        break
            if start_row: break
            
        # 2. Fallback if tag not found: Find 'Stt' and add offset
        if not start_row: 
            for row in ws.iter_rows(min_row=1, max_row=30):
                for cell in row:
                    # We do NOT skip merged cells for searching content, as the header might be merged.
                    if cell.value and isinstance(cell.value, str):
                        if "STT" == cell.value.strip().upper():
                            # User feedback: Header is merged, data starts 2 rows below STT
                            start_row = cell.row + 2
                            break
                if start_row: break
        
        if not start_row: start_row = 15 # Fallback
        
        # CLEAR existing rows below header to remove template artifacts (like {{i.ten_loi}})
        # Let's clear next 10 rows to be safe
        for r in range(start_row, start_row + 20):
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

        if not details_df.empty:
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            c_idx = start_row
            for i, row in details_df.iterrows():
                # Helper to safely set value
                def safe_set(r, c, val, align):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell): return # Skip if merged
                    cell.value = val
                    cell.border = border
                    cell.alignment = align

                # 1. STT (Col 1)
                safe_set(c_idx, 1, i + 1, align_center)
                
                # 2. Ten Loi (Col 2)
                safe_set(c_idx, 2, row.get('defect_name', ''), align_left)
                
                # Cols 3, 4, 5, 6 (Forms of processing/Place) -> Leave empty for manual check/input
                # Do NOT fill with 0
                
                # 5. SL Hong (Col 7) -> Should be empty per request (was Qty Assigned)
                safe_set(c_idx, 7, "", align_center)
                
                # Other columns left empty per instruction to fix positions
                
                c_idx += 1

        # --- 4. SAVE ---
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer
        
    except Exception as e:
        st.error(f"Lỗi xuất Excel: {e}")
        return None
